from Matrix import *
import time
import openpyxl

n = [1024, 2048, 4096, 8192]

wb = openpyxl.load_workbook('../../analysis/data.xlsx')
print("workbook", wb.sheetnames, "loaded")
ws = wb['multiplication']

i = 1
for d in n:
    print("test for", d, "by", d, "matrix starts")
    print("Matrix generation starts")
    m1 = random_matrix_gen(d)
    print("Matrix 1 generated")
    m2 = random_matrix_gen(d)
    print("Matrix 2 generated")

    ws['C' + str(i)] = d

    time1 = time.time()
    c = strassen_multiply(m1, m2, 128)
    time2 = time.time()
    print("The run time of Strassen's method is", time2 - time1)
    ws['D' + str(i)] = time2 - time1
    print("==========")

    wb.save('../analysis/data.xlsx')
    i += 1

print("test finished")
