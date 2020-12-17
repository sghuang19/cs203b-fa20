from Matrix import *
import time
import openpyxl

n = 512
s = 1
e = 64
r = 2

print("Matrix generation starts")
m1 = random_matrix_gen(n)
print("Matrix 1 generated")
m2 = random_matrix_gen(n)
print("Matrix 2 generated")

wb = openpyxl.load_workbook('../analysis/data.xlsx')
print("workbook", wb.sheetnames, "loaded")
ws = wb['crossover_point']

for i in range(s, e + 1):
    strassen = 0
    square = 0
    ws['A' + str(i)] = i
    print("The recursion point is set to be ", i)

    for j in range(1, r + 1):
        print("test ", j)

        time1 = time.time()
        c = strassen_multiply(m1, m2, i)
        time2 = time.time()
        print("The run time of Strassen's method is", time2 - time1)
        strassen = time2 - time1 + strassen

        time1 = time.time()
        c1 = square_matrix_multiply(m1, m2)
        time2 = time.time()
        print("The run time of brutal method is", time2 - time1)
        square = time2 - time1 + square

    print("average strassen", strassen / r)
    ws['B' + str(i)] = strassen / r
    print("average square", square / r)
    ws['C' + str(i)] = square / r
    print("==========")

    wb.save('data.xlsx')

print("benchmark finished")
