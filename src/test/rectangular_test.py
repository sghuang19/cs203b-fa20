from Matrix import *
import time
import openpyxl

m = 256
n = 256

wb = openpyxl.load_workbook('../analysis/data.xlsx')
print("workbook", wb.sheetnames, "loaded")
ws = wb['rectangular']

i = 1
for n in range(m, 3 * m + 1):
    print("test for", m, 'by', n, 'matrix starts')
    ws['A' + str(i)] = n

    print("Matrix generation starts")
    m1 = random_matrix_gen(m, n)
    print("Matrix 1 generated")
    m2 = random_matrix_gen(n, m)
    print("Matrix 2 generated")

    time1 = time.time()
    c = strassen_multiply(m1, m2)
    time2 = time.time()
    print("The run time of Strassen's method is", time2 - time1)
    ws['B' + str(i)] = time2 - time1

    time1 = time.time()
    c1 = square_matrix_multiply(m1, m2)
    time2 = time.time()
    print("The run time of brutal method is", time2 - time1)
    ws['C' + str(i)] = time2 - time1

    print("==========")

    wb.save('../analysis/data.xlsx')
    n += 16
    i += 1

print("benchmark finished")
