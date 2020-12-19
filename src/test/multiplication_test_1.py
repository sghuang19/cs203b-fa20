from Matrix import *
import time
import openpyxl

# s = 16
s = 656
# s = 1024
# e = 2048
# e = 1024
e = 1200
# step = 256
step = 16

wb = openpyxl.load_workbook('../../analysis/data.xlsx')
print("workbook", wb.sheetnames, "loaded")
ws = wb['multiplication']

n = s
i = 1

while n <= e:
    print("test for", n, "by", n, "matrix starts")
    print("Matrix generation starts")
    m1 = random_matrix_gen(n)
    print("Matrix 1 generated")
    m2 = random_matrix_gen(n)
    print("Matrix 2 generated")

    ws['G' + str(i)] = n

    time1 = time.time()
    # c = strassen_multiply(m1, m2, 128)
    c = square_matrix_multiply(m1, m2)
    time2 = time.time()
    print("The run time of Strassen's method is", time2 - time1)
    ws['H' + str(i)] = time2 - time1
    print("==========")

    wb.save('../analysis/data.xlsx')

    n += step
    i += 1

print("benchmark finished")
