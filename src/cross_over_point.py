from matrix import Matrix, square_matrix_multiply, strassen_multiply
import time
import random
import openpyxl

n = 20
s = 35
e = 45

elements = []
for i in range(n * n):
    elements.append(random.uniform(-1, 1))
m1 = Matrix(elements, n, n)

elements = []
for i in range(n * n):
    elements.append(random.uniform(-1, 1))
m2 = Matrix(elements, n, n)

wb = openpyxl.load_workbook('data.xlsx')
print("workbook", wb.sheetnames, "loaded")
# ws = workbook['test']
ws = wb.active

for i in range(35, 45):
    ws['A' + str(i)] = i
    print("==========")
    print("The recursion point is set to be ", i)
    time1 = time.time()
    c = strassen_multiply(m1, m2)
    time2 = time.time()
    print("The run time of Strassen's method is", time2 - time1)

    time1 = time.time()
    c1 = square_matrix_multiply(m1, m2)
    time2 = time.time()
    print("The run time of brutal method is", time2 - time1)

    print("==========")

wb.save('data.xlsx')