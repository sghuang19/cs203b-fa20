from matrix import strassen_multiply, random_matrix_gen
import time
import openpyxl

s = 16
e = 4096
step = 16

wb = openpyxl.load_workbook('data.xlsx')
print("workbook", wb.sheetnames, "loaded")
ws = wb['multiplication']

n = s
i = 1

while n <= e:
    print("test for", n, "by", n, "matrix starts")
    print("Matrix generation starts")
    m1 = random_matrix_gen(n)
    print("Matrix 1 generated")
    m2 = random_matrix_gen(n)
    print("Matrix 2 generated")

    ws['A' + str(i)] = n

    time1 = time.time()
    c = strassen_multiply(m1, m2, 128)
    time2 = time.time()
    print("The run time of Strassen's method is", time2 - time1)
    ws['B' + str(i)] = time2 - time1
    print("==========")

    wb.save('data.xlsx')

    n += step
    i += 1

print("benchmark finished")
