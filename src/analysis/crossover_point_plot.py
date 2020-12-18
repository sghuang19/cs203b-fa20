from pylab import *
import numpy as np
import matplotlib.pyplot as plt
import openpyxl

X = np.arange(1, 64 + 1)
strassen = np.zeros(64)
square = np.zeros(64)

wb = openpyxl.load_workbook("data.xlsx")
ws = wb['crossover_point']
for i in range(0, 64):
    strassen[i] = ws['B' + str(i + 1)].value
    square[i] = ws['C' + str(i + 1)].value

plt.figure(figsize=(6.4, 4.8), dpi=600)
plt.plot(X, strassen, label='Strassen\'s method')
plt.plot(X, square, label='standard method')
plt.title('The runtime of matrix multiplication with varied recursion points'
          '\n(matrix size is set to be 512 by 512)')
plt.xlabel('Recursion point')
plt.ylabel('Runtime/s')
plt.legend(loc='upper right')

strassen_min = np.argmin(strassen)
strassen_max = np.argmax(strassen)
strassen_show_min = '[' + str(strassen_min) + ' {:.4f}]'.format(strassen[strassen_min])
strassen_show_max = '[' + str(strassen_max) + ' {:.4f}]'.format(strassen[strassen_max])
plt.annotate(strassen_show_min, xy=(strassen_min, strassen[strassen_min]),
             xytext=(strassen_min, strassen[strassen_min]))
plt.annotate(strassen_show_max, xy=(strassen_max, strassen[strassen_max]),
             xytext=(strassen_max, strassen[strassen_max]))

square_min = np.argmin(square)
square_max = np.argmax(square)
square_show_min = '[' + str(square_min) + ' {:.4f}]'.format(square[square_min])
square_show_max = '[' + str(square_max) + ' {:.4f}]'.format(square[square_max])
plt.annotate(square_show_min, xy=(square_min, square[square_min]),
             xytext=(square_min, square[square_min]))
plt.annotate(square_show_max, xy=(square_max, square[square_max]),
             xytext=(square_max, square[square_max]))

plt.show()

difference = np.zeros(64)
for i in range(64):
    difference[i] = (square[i] - strassen[i]) / square[i] * 100
plt.figure(figsize=(6.4, 4.8), dpi=600)
plt.plot(X, difference)
plt.title('The difference between runtime of standard and Strassen\'s method\n'
          'with varied recursion points '
          '(matrix size is set to be 512 by 512)')
plt.xlabel('Recursion point')
plt.ylabel('Difference/percent')

min = np.argmin(difference)
max = np.argmax(difference)
show_min = '[' + str(min) + ' {:.4f}]'.format(difference[min])
show_max = '[' + str(max) + ' {:.4f}]'.format(difference[max])
plt.annotate(show_min, xy=(min, difference[min]),
             xytext=(min, difference[min]))
plt.annotate(show_max, xy=(max, difference[max]),
             xytext=(max, difference[max]))

plt.show()
