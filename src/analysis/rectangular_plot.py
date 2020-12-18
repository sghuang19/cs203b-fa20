from pylab import *
import numpy as np
import matplotlib.pyplot as plt
import openpyxl

strassen = np.zeros(9)
standard = np.zeros(9)
r = np.zeros(9)

wb = openpyxl.load_workbook("data.xlsx")
ws = wb['rectangular']

for i in range(0, 9):
    r[i] = ws['A' + str(i + 1)].value / 256
    strassen[i] = ws['B' + str(i + 1)].value
    standard[i] = ws['C' + str(i + 1)].value

c_strassen = np.polyfit(r, strassen, 1)
print(c_strassen)
f_strassen = np.polyval(c_strassen, r)

c_standard = np.polyfit(r, standard, 1)
print(c_standard)
f_standard = np.polyval(c_standard, r)

plt.figure(figsize=(6.4, 4.8), dpi=600)
plt.plot(r, strassen, 'x', label='runtime of Strassen\'s method')
plt.plot(r, f_strassen, label='polynomial fitted runtime of Strassen\'s method')
plt.plot(r, standard, 'x', label='runtime of standard method')
plt.plot(r, f_standard, label='polynomial fitted runtime of standard method')
plt.grid()
plt.title('The runtime of matrix multiplication with varied shape\n'
          'recursion point is set to be 128')
plt.xlabel('Matrix shape (Ratio of long edge over short edge)')
plt.ylabel('Runtime/s')
plt.legend(loc='upper left')

plt.show()
