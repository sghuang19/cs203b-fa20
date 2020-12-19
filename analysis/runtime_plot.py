from pylab import *
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from scipy.optimize import curve_fit

x = 80
strassen = np.zeros(x)
square = np.zeros(x)
n = np.zeros(x)

wb = openpyxl.load_workbook("data.xlsx")
ws = wb['multiplication']
for i in range(0, x):
    n[i] = ws['A' + str(i + 1)].value
    strassen[i] = ws['B' + str(i + 1)].value
    square[i] = ws['D' + str(i + 1)].value


# n = np.append(n, [1024, 2048])
# t = np.append(t, [ws['D1'].value, ws['D2'].value])


def func(xval, c1, c2, c3, c4):
    return c1 * xval ** 2.81 + c2 * xval ** 2 + c3 * xval + c4


popt, pcov = curve_fit(func, n, strassen)
c1 = popt[0]
c2 = popt[1]
c3 = popt[2]
c4 = popt[3]
yvals = func(n, c1, c2, c3, c4)
print(c1, c2, c3, c4)

c = np.polyfit(n, square, 3)
print(c)
f = np.polyval(c, n)

# plt.figure(figsize=(6.4, 4.8), dpi=600)
# # plt.subplot(1, 2, 1)
# plt.plot(n, t, 'x', label='runtime of Strassen\'s method')
# plt.plot(n, f, label=r'$n^3$ order polynomial fitted runtime')
# plt.grid()
# plt.title('The runtime of matrix multiplication with varied matrix size\n'
#           'recursion point is set to be 128')
# plt.xlabel('Matrix size (n by n)')
# plt.ylabel('Runtime/s')
# plt.legend(loc='upper left')

# figure()
plt.figure(figsize=(6.4, 4.8), dpi=600)
plt.grid()
plt.title('The runtime of matrix multiplication with varied matrix size\n'
          'recursion point is set to be 128')
plt.xlabel('Matrix size (n by n)')
plt.ylabel('Runtime/s')
# plt.subplot(1, 2, 2)

plt.plot(n, strassen, 'x', label='runtime of Strassen\'s method')
plt.plot(n, yvals, '--', label=r'$n^{2.81}$ polynomial fitted runtime for Strassen' + '\'s method')

plt.plot(n, square, '.', label='runtime of standard method')
plt.plot(n, f, '--', label=r'$n^3$ polynomial fitted runtime of standard method')

plt.legend(loc='upper left')

# plt.subplots_adjust()
plt.show()
